"""
Генератор Fin_Model_Microgreen_Uzbekistan.xlsx.

Цифры не набираются здесь — они приходят из `model.py`. Запуск:

    C:/Users/noteb/AppData/Local/Programs/Python/Python310/python.exe -X utf8 \
        doc/create_fin_model.py

openpyxl есть только в системном Python 3.10; в .venv/.test_venv его нет.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
import model as M  # noqa: E402

OUT = Path(__file__).resolve().parent / "ready" / "Fin_Model_Microgreen_Uzbekistan.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="1F6F43")
HEAD_FILL = PatternFill("solid", fgColor="D9EAD3")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(bold=True, size=13, color="FFFFFF")
HEAD_FONT = Font(bold=True, size=10)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUM = "#,##0"


def _title(ws, text: str, width: int) -> None:
    ws.cell(row=1, column=1, value=text).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    ws.row_dimensions[1].height = 22


def _header(ws, row: int, labels: list[str], first: str, total: str | None) -> None:
    ws.cell(row=row, column=1, value=first).font = HEAD_FONT
    ws.cell(row=row, column=1).fill = HEAD_FILL
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=2 + i, value=label)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center")
    if total:
        cell = ws.cell(row=row, column=2 + len(labels), value=total)
        cell.font = HEAD_FONT
        cell.fill = TOTAL_FILL
        cell.alignment = Alignment(horizontal="center")


def _series(ws, row: int, name: str, values, total=None, bold=False, fmt=NUM) -> int:
    cell = ws.cell(row=row, column=1, value=name)
    if bold:
        cell.font = BOLD
    for i, value in enumerate(values):
        c = ws.cell(row=row, column=2 + i, value=value)
        c.number_format = fmt
        c.border = BORDER
        if bold:
            c.font = BOLD
    if total is not None:
        c = ws.cell(row=row, column=2 + len(values), value=total)
        c.number_format = fmt
        c.fill = TOTAL_FILL
        c.font = BOLD
    return row + 1


def _widths(ws, first: int, rest: int, count: int) -> None:
    ws.column_dimensions["A"].width = first
    for i in range(count + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = rest


def build_workbook(m: dict) -> Workbook:
    labels = m["labels"]
    n = len(labels)
    wb = Workbook()

    # ── 1. Клиенты ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "1. Mijozlar"
    _title(ws, "1. MIJOZLAR — baza va o'sish", n + 2)
    ws.cell(row=2, column=1,
            value=f"M1 — bugungi FAKT: {m['restaurants'][0]} restoran + "
                  f"{m['families'][0]} oila (obuna) + ~{m['retail_purchases'][0]} chakana "
                  f"xarid = {m['revenue'][0]:,} so'm/oy. O'sish shu nuqtadan.")
    _header(ws, 4, labels, "Ko'rsatkich", "Oxirgi oy")
    row = 5
    row = _series(ws, row, "Restoranlar (B2B)", m["restaurants"], m["restaurants"][-1])
    row = _series(ws, row, "Oilalar (obuna)", m["families"], m["families"][-1])
    row = _series(ws, row, "DOIMIY MIJOZLAR", m["regulars"], m["regulars"][-1], bold=True)
    row = _series(ws, row, "Chakana xaridlar (dona/oy)", m["retail_purchases"],
                  m["retail_purchases"][-1])
    row = _series(ws, row, "Yangi mijozlar", m["new_customers"], sum(m["new_customers"]))
    row = _series(ws, row, "Churn (oylik, %)", [M.MONTHLY_CHURN * 100] * n,
                  M.MONTHLY_CHURN * 100, fmt="0.0")
    _widths(ws, 26, 11, n)

    # ── 2. Тарифы ────────────────────────────────────────────────────────
    ws = wb.create_sheet("2. Tariflar")
    _title(ws, "2. TARIFLAR — modelda ishlatilgan narxlar", 4)
    _header(ws, 3, ["Narx (UZS)", "Narx ($)", "Izoh"], "Kanal / Mahsulot", None)
    rows = [
        ("Restoran — bitta yetkazish cheki", M.CHECK_RESTAURANT,
         f"Oyiga ~{M.DELIVERIES_PER_MONTH} marta yetkazamiz → "
         f"{M.CHECK_RESTAURANT * M.DELIVERIES_PER_MONTH:,} so'm/oy"),
        ("Oila obunasi (oylik)", M.PRICE_FAMILY,
         "Yagona haqiqiy obuna kanali"),
        ("Chakana savdo — o'rtacha chek", M.CHECK_RETAIL,
         "Mijozlar donalab oladi, eng katta oqim"),
        ("Premium qulupnay (1 kg)", M.STRAWBERRY_PRICE_KG,
         "Yangi ferma mahsuloti, asosan chakanaga"),
    ]
    for i, (name, price, note) in enumerate(rows):
        r = 4 + i
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=price).number_format = NUM
        ws.cell(row=r, column=3, value=round(price / M.USD_RATE, 2)).number_format = "0.00"
        ws.cell(row=r, column=4, value=note)
    ws.cell(row=8, column=1, value="Ferma quvvati (fizik chegara):").font = BOLD
    ws.cell(row=9, column=1, value=f"{M.STRAWBERRY_PLANTS} tup × "
                                   f"{M.KG_PER_PLANT_YEAR} kg/yil = "
                                   f"{M.STRAWBERRY_PLANTS * M.KG_PER_PLANT_YEAR:,.0f} kg/yil")
    ws.cell(row=10, column=1,
            value=f"= {m['strawberry_capacity']:,} so'm/oy — bundan ortiq 60 m² dan "
                  f"olib bo'lmaydi. Kengaytirish keyingi bosqich.")
    ws.column_dimensions["A"].width = 38
    for col in "BCD":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["D"].width = 52

    # ── 3. Доход ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("3. Daromad")
    _title(ws, "3. DAROMAD", n + 2)
    _header(ws, 3, labels, "Ko'rsatkich (UZS)", "JAMI")
    row = 4
    row = _series(ws, row, "Mijozlar to'lovi (QQS bilan)", m["revenue_gross"],
                  sum(m["revenue_gross"]))
    row = _series(ws, row, f"QQS ({M.VAT_RATE:.0%})", m["vat_payable"],
                  sum(m["vat_payable"]))
    row += 1
    row = _series(ws, row, "Restoranlar (buyurtma)", m["b2b"], sum(m["b2b"]))
    row = _series(ws, row, "Oila obunalari", m["subscriptions"], sum(m["subscriptions"]))
    row = _series(ws, row, "Chakana savdo", m["retail"], sum(m["retail"]))
    row = _series(ws, row, "Qulupnay", m["strawberry"], sum(m["strawberry"]))
    row = _series(ws, row, "JAMI DAROMAD (QQSsiz)", m["revenue"], sum(m["revenue"]),
                  bold=True)
    row += 1
    row = _series(ws, row, "MRR (takrorlanuvchi)", m["mrr"], m["mrr"][-1])
    row = _series(ws, row, "ARR (yillik)", m["arr"], m["arr"][-1])
    row = _series(ws, row, "ARPU", m["arpu"], m["arpu"][-1])
    row = _series(ws, row, "Daromad ($)", [round(v / M.USD_RATE) for v in m["revenue"]],
                  round(sum(m["revenue"]) / M.USD_RATE))
    _widths(ws, 26, 13, n)

    # ── 4. Расходы ───────────────────────────────────────────────────────
    ws = wb.create_sheet("4. Xarajatlar")
    _title(ws, "4. XARAJATLAR", n + 2)
    _header(ws, 3, labels, "Xarajat turi (UZS)", "JAMI")
    row = 4
    for name, values in m["opex_rows"].items():
        row = _series(ws, row, name, values, sum(values))
    row = _series(ws, row, "JAMI OPEX", m["opex_total"], sum(m["opex_total"]), bold=True)
    row += 1
    row = _series(ws, row, "COGS (tannarx)", m["cogs"], sum(m["cogs"]))
    row = _series(ws, row, "Brutto foyda", m["gross"], sum(m["gross"]), bold=True)
    row = _series(ws, row, "Brutto marja (%)",
                  [round(v * 100, 1) for v in m["gross_margin"]],
                  round(sum(m["gross"]) / sum(m["revenue"]) * 100, 1), fmt="0.0")
    _widths(ws, 26, 13, n)

    # ── 5. Оборотные средства ────────────────────────────────────────────
    ws = wb.create_sheet("5. Aylanma mablag'")
    _title(ws, "5. AYLANMA MABLAG' — investitsiyaning to'liq taqsimoti", n + 2)
    _header(ws, 3, labels, "Ko'rsatkich (UZS)", "JAMI")
    row = 4
    row = _series(ws, row, "Investitsiya kirimi", m["investment"], sum(m["investment"]))
    row = _series(ws, row, "Mijozlardan tushum (QQS bilan)", m["revenue_gross"],
                  sum(m["revenue_gross"]))
    row = _series(ws, row, "JAMI KIRIM", m["inflow"], sum(m["inflow"]), bold=True)
    row += 1
    row = _series(ws, row, "CAPEX (uskunalar)", m["capex"], sum(m["capex"]))
    row = _series(ws, row, "Bino tayyorlash", m["building_prep"], sum(m["building_prep"]))
    row = _series(ws, row, "COGS (xarid, tannarx)", m["cogs"], sum(m["cogs"]))
    row = _series(ws, row, "OPEX", m["opex_total"], sum(m["opex_total"]))
    row = _series(ws, row, "QQS to'lovi", m["vat_payable"], sum(m["vat_payable"]))
    row = _series(ws, row, "Foyda solig'i", m["tax"], sum(m["tax"]))
    row = _series(ws, row, "JAMI CHIQIM", m["outflow"], sum(m["outflow"]), bold=True)
    row += 1
    row = _series(ws, row, "SOF PUL OQIMI", m["net_cf"], sum(m["net_cf"]), bold=True)
    row = _series(ws, row, "KASSA QOLDIG'I", m["cash"], m["cash"][-1], bold=True)
    _widths(ws, 26, 13, n)

    # ── 6. Прибыль ───────────────────────────────────────────────────────
    ws = wb.create_sheet("6. Sof foyda")
    _title(ws, "6. SOF FOYDA (P&L)", n + 2)
    _header(ws, 3, labels, "P&L (UZS)", "JAMI")
    row = 4
    row = _series(ws, row, "Jami daromad (QQSsiz)", m["revenue"], sum(m["revenue"]))
    row = _series(ws, row, "COGS", m["cogs"], sum(m["cogs"]))
    row = _series(ws, row, "BRUTTO FOYDA", m["gross"], sum(m["gross"]), bold=True)
    row = _series(ws, row, "OPEX", m["opex_total"], sum(m["opex_total"]))
    row = _series(ws, row, "EBITDA", m["ebitda"], sum(m["ebitda"]), bold=True)
    row = _series(ws, row, "Amortizatsiya", m["depreciation"], sum(m["depreciation"]))
    row = _series(ws, row, "Operatsion foyda (EBIT)", m["ebit"], sum(m["ebit"]))
    row = _series(ws, row, f"Foyda solig'i ({M.PROFIT_TAX_RATE:.0%})", m["tax"],
                  sum(m["tax"]))
    row = _series(ws, row, "SOF FOYDA", m["net"], sum(m["net"]), bold=True)
    row = _series(ws, row, "Sof marja (%)",
                  [round(m["net"][i] / m["revenue"][i] * 100, 1) for i in range(n)],
                  round(sum(m["net"]) / sum(m["revenue"]) * 100, 1), fmt="0.0")
    row += 1
    row = _series(ws, row, "Kumulyativ sof foyda", m["cumulative"], m["cumulative"][-1])
    _widths(ws, 26, 13, n)

    # ── 7. Startup KPI ───────────────────────────────────────────────────
    ws = wb.create_sheet("7. Startup KPI")
    _title(ws, "7. STARTAP KO'RSATKICHLARI", n + 2)
    _header(ws, 3, labels, "Metrika", "O'rtacha")
    row = 4
    row = _series(ws, row, "MRR", m["mrr"], m["mrr"][-1])
    row = _series(ws, row, "ARR", m["arr"], m["arr"][-1])
    row = _series(ws, row, "ARPU", m["arpu"], round(sum(m["arpu"]) / n))
    cac_display = [v if v is not None else "—" for v in m["cac"]]
    valid_cac = [v for v in m["cac"] if v]
    row = _series(ws, row, "CAC", cac_display, round(sum(valid_cac) / len(valid_cac)))
    row = _series(ws, row, "LTV", m["ltv"], round(sum(m["ltv"]) / n))
    ratio = [round(m["ltv"][i] / m["cac"][i], 1) if m["cac"][i] else "—" for i in range(n)]
    valid_ratio = [v for v in ratio if v != "—"]
    row = _series(ws, row, "LTV/CAC", ratio,
                  round(sum(valid_ratio) / len(valid_ratio), 1), fmt="0.0")
    row += 1
    ws.cell(row=row, column=1, value="LTV = ARPU × brutto marja × mijoz umri").font = BOLD
    row += 1
    ws.cell(row=row, column=1,
            value=f"Mijoz umri = min(1/churn, {M.MONTHS} oy) = "
                  f"{min(1 / M.MONTHLY_CHURN, M.MONTHS):.0f} oy — model gorizontidan "
                  f"uzoqroq daromad va'da qilinmaydi")
    row += 1
    ws.cell(row=row, column=1,
            value=f"CAC = (marketing + direktor ish haqining {M.SALES_SALARY_SHARE:.0%}) "
                  f"/ yangi mijozlar — degustatsiyalarni direktor o'zi o'tkazadi")
    _widths(ws, 26, 13, n)

    # ── XULOSA ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("XULOSA")
    _title(ws, "MICROGREEN AND STRAWBERRY — FIN MODEL XULOSA", 3)
    items = [
        ("INVESTITSIYA", "", ""),
        ("So'ralayotgan summa", f"${M.INVESTMENT_USD:,}", f"{M.INVESTMENT_UZS:,} UZS"),
        ("  — moddalar", f"${sum(M.BUDGET_USD.values()):,}",
         f"{sum(M.BUDGET_USD.values()) * M.USD_RATE:,} UZS"),
        ("  — zaxira", f"${M.RESERVE_USD:,}", f"{M.RESERVE_USD * M.USD_RATE:,} UZS"),
        ("Valyuta kursi", f"{M.USD_RATE:,} UZS/$", ""),
        ("CAPEX (amortizatsiya bazasi)", f"{m['capex_uzs']:,} UZS",
         f"{M.DEPRECIATION_MONTHS} oyda amortizatsiya"),
        ("", "", ""),
        ("DAVR", f"{m['labels'][0]} — {m['labels'][-1]}", f"{M.MONTHS} oy"),
        ("Boshlang'ich baza (M1)", f"{m['revenue_gross'][0]:,} UZS/oy",
         f"{m['restaurants'][0]} restoran + {m['families'][0]} oila obuna + "
         f"~{m['retail_purchases'][0]} chakana xarid — FAKT, QQS bilan"),
        ("  — QQSsiz daromad", f"{m['revenue'][0]:,} UZS/oy",
         "P&L shu summadan hisoblanadi"),
        ("Soliq rejimi", f"umumiy: QQS {M.VAT_RATE:.0%}",
         f"foyda solig'i {M.PROFIT_TAX_RATE:.0%} — aylanma solig'i emas"),
        ("Asoschi allaqachon kiritgan", f"${M.FOUNDER_INVESTED_USD:,}",
         "shaxsiy mablag', tashqi investitsiyasiz"),
        ("", "", ""),
        ("DAROMAD", "", ""),
        ("Jami daromad", f"{sum(m['revenue']):,} UZS",
         f"${round(sum(m['revenue']) / M.USD_RATE):,}"),
        ("Oxirgi oy daromadi", f"{m['revenue'][-1]:,} UZS",
         f"${round(m['revenue'][-1] / M.USD_RATE):,}"),
        ("MRR (oxirgi oy)", f"{m['mrr'][-1]:,} UZS", ""),
        ("ARR (oxirgi oy)", f"{m['arr'][-1]:,} UZS",
         f"${round(m['arr'][-1] / M.USD_RATE):,}"),
        ("", "", ""),
        ("FOYDA", "", ""),
        ("Jami sof foyda", f"{sum(m['net']):,} UZS",
         f"${round(sum(m['net']) / M.USD_RATE):,}"),
        ("EBITDA marja (o'rtacha)",
         f"{sum(m['ebitda']) / sum(m['revenue']) * 100:.1f}%", "soliqqacha"),
        ("Break-even", m["labels"][m["break_even_index"]],
         "kompaniya bugun ham foydali — investitsiya o'sish uchun"),
        ("ROI (24 oy)", f"{sum(m['net']) / M.INVESTMENT_UZS * 100:.0f}%",
         "transh kelgan kundan hisoblanadi"),
        ("", "", ""),
        ("MIJOZLAR", "", ""),
        ("Doimiy mijozlar (M24)", f"{m['regulars'][-1]}",
         f"{m['restaurants'][-1]} restoran + {m['families'][-1]} oila obuna"),
        ("Chakana xaridlar (M24)", f"{m['retail_purchases'][-1]}/oy",
         f"o'rtacha chek {M.CHECK_RETAIL:,} so'm"),
        ("Churn (oylik)", f"{M.MONTHLY_CHURN * 100:.1f}%", ""),
        ("LTV/CAC (o'rtacha)",
         f"{sum(valid_ratio) / len(valid_ratio):.1f}x", "formula 7-varaqda"),
        ("", "", ""),
        ("KASSA", "", ""),
        ("Kassa qoldig'i (M24)", f"{m['cash'][-1]:,} UZS",
         f"${round(m['cash'][-1] / M.USD_RATE):,}"),
    ]
    for i, (name, value, note) in enumerate(items):
        r = 3 + i
        cell = ws.cell(row=r, column=1, value=name)
        if value == "" and name:
            cell.font = BOLD
            cell.fill = HEAD_FILL
        ws.cell(row=r, column=2, value=value)
        ws.cell(row=r, column=3, value=note)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 48

    return wb


def main() -> int:
    m = M.build()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(m).save(OUT)
    print(f"[OK] {OUT}")
    print(f"     davr: {m['labels'][0]} — {m['labels'][-1]}")
    print(f"     M1 daromad: {m['revenue'][0]:,} (fakt)")
    print(f"     M24 daromad: {m['revenue'][-1]:,}")
    print(f"     24 oylik sof foyda: {sum(m['net']):,}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
